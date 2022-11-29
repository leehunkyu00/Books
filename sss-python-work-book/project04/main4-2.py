from openpyxl import load_workbook
from datetime import date

# Read quote file
quote_filepath = r"견적서_샘플.xlsx"

quote_wb = load_workbook(quote_filepath, data_only=True)
quote_ws = quote_wb.active

## read one cel
quote_receiver = quote_ws['A4'].value
quote_price = quote_ws['X25'].value
quote_tax = quote_ws['X26'].value
quote_total = quote_ws['X27'].value

print('quote_receiver', quote_receiver)
print('quote_price', quote_price)
print('quote_tax', quote_tax)
print('quote_total',  quote_total)

name_list = []
count_list = []
price_list = []
total_list = []

for data in quote_ws['C13':'X24']:
    for cell in data:
        if cell.column == ord('C')-64:
            if cell.value is not None:
                name_list.append(cell.value)
        if cell.column == ord('R')-64:
            if cell.value is not None:
                count_list.append(cell.value)
        if cell.column == ord('T')-64:
            if cell.value is not None:
                price_list.append(cell.value)
        if cell.column == ord('X')-64:
            if cell.value is not None:
                total_list.append(cell.value)


print('name_list', name_list)
print('count_list', count_list)
print('price_list', price_list)
print('total_list', total_list)

# make transaction file
transaction_filepath = r"거래명세표_샘플.xlsx"

transaction_wb = load_workbook(transaction_filepath, data_only=False)
transaction_ws = transaction_wb.active

today = date.today()
transaction_ws['C4'].value = today.year
transaction_ws['E4'].value = today.month
transaction_ws['G4'].value = today.day

transaction_ws['C6'].value = quote_receiver

for i in range(len(name_list)):
    transaction_ws.cell(row=i+10, column=ord('B')-64).value = name_list[i]
    transaction_ws.cell(row=i+10, column=ord('G')-64).value = count_list[i]
    transaction_ws.cell(row=i+10, column=ord('I')-64).value = price_list[i]

transaction_wb.save('거래명세표_'+quote_receiver+".xlsx")


